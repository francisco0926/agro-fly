
import openpyxl
import numpy as np
import pandas as pd
from fpdf import FPDF
import os
import streamlit as st
import io
import zipfile
import math

# 1. CONFIGURACIÃ“N INICIAL
st.set_page_config(page_title="AgroReport Pro", layout="centered", page_icon="ğŸŒ¾")

# --- FUNCIONES DE APOYO ---
def convertir_a_segundos(tiempo_str):
    try:
        minutos, segundos = map(int, str(tiempo_str).split(':'))
        return (minutos * 60) + segundos
    except: return 0

def formatear_tiempo(segundos_totales):
    horas = segundos_totales // 3600
    minutos = (segundos_totales % 3600) // 60
    segundos = segundos_totales % 60
    return f"{int(horas):02d}:{int(minutos):02d}:{int(segundos):02d}"

# --- CLASE PDF ---
class PDF_Decorado(FPDF):
    def __init__(self, nombre_empresa="AGRO REPORT"):
        super().__init__()
        self.nombre_empresa = nombre_empresa # Guardamos el nombre aquÃ­
    def dibujar_logo_drone(self, x, y):
        # Color dorado/trigo para la espiga
        self.set_draw_color(255, 215, 0) # Dorado
        self.set_fill_color(255, 215, 0)
        self.set_line_width(0.6)

        # 1. Tallo central (una lÃ­nea inclinada como la espiga ğŸŒ¾)
        self.line(x + 5, y + 15, x + 5, y + 2) 

        # 2. Granos de la espiga (pequeÃ±as elipses a los costados)
        # Lado izquierdo
        self.ellipse(x + 2, y + 4, 3, 2, 'F')
        self.ellipse(x + 1, y + 8, 3, 2, 'F')
        self.ellipse(x + 2, y + 12, 3, 2, 'F')
    
        # Lado derecho
        self.ellipse(x + 6, y + 6, 3, 2, 'F')
        self.ellipse(x + 7, y + 10, 3, 2, 'F')
        self.ellipse(x + 6, y + 14, 3, 2, 'F')

        # 3. Un pequeÃ±o cÃ­rculo en la punta
        self.ellipse(x + 4, y, 2.5, 3.5, 'F')


    def header(self):
        self.set_fill_color(0, 51, 102)
        self.rect(0, 0, 210, 40, 'F')
        self.dibujar_logo_drone(170, 12)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 20)
        #self.cell(190, 15, '  AGRO REPORT', 0, 1, 'L') 
        self.cell(190, 15, f'  {self.nombre_empresa}', 0, 1, 'L')
        self.ln(20)

    def footer(self):
        self.set_y(-20)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'PÃ¡gina {self.page_no()}', 0, 0, 'C')

# --- PROCESADOR DE DATOS ---
def procesar_datos_informe(df_subido):
    df = df_subido.copy()
    df.columns = df.columns.astype(str).str.strip()
    df['fecha_simple'] = df['Flight time'].astype(str).str[:10]
    df['area_num'] = pd.to_numeric(df['Sprayed area'], errors='coerce').fillna(0)
    df['segundos_vuelo'] = df['Flight duration(min:sec)'].apply(convertir_a_segundos)
    df['insumo_num'] = pd.to_numeric(df['Total Amount(L/Kg)'], errors='coerce').fillna(0)

    informe = df.groupby(['fecha_simple', 'Location']).agg({
        'area_num': 'sum', 'insumo_num': 'sum', 'segundos_vuelo': 'sum', 'Flight time': 'count'
    }).reset_index()

    def corregir(row):
        mins = row['segundos_vuelo'] / 60
        ha = row['area_num']
        return ha / 10 if (mins > 0 and 1 < (ha/mins) < 10) else ha
    
    informe['Area Final'] = informe.apply(corregir, axis=1)
    # Creamos una etiqueta amigable para el usuario
    informe['etiqueta'] = informe['fecha_simple'] + " | " + informe['Location']
    return informe

# --- GENERADOR DE ZIP FILTRADO ---
def generar_zip_seleccionado(informe_filtrado):
    # Obtenemos el nombre de la sesiÃ³n, si no existe usamos "AGROFLY"
    nombre_personalizado = st.session_state.get('nombre_empresa', 'AGROFLY')
    buffer_zip = io.BytesIO()
    with zipfile.ZipFile(buffer_zip, "w") as zf:
        for i, fila in informe_filtrado.iterrows():
            pdf = PDF_Decorado(nombre_empresa=nombre_personalizado)
            pdf.set_margins(10, 10, 10)
            pdf.add_page()
            
            pdf.set_text_color(0, 51, 102)
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(190, 10, f"ORDEN DE TRABAJO: {fila['fecha_simple']}", "B", 1, 'L')
            pdf.ln(10)

            def agregar_fila_dato(pdf_obj, label, valor, unidad=""):
                texto_completo = f" {valor} {unidad}"
                ancho_texto = pdf_obj.get_string_width(texto_completo)
                lineas = math.ceil(ancho_texto / 125)
                if lineas < 1: lineas = 1
                altura_total = lineas * 12
                y_actual = pdf_obj.get_y()

                pdf_obj.set_fill_color(240, 245, 255)
                pdf_obj.set_font('Arial', 'B', 11)
                pdf_obj.cell(60, altura_total, f" {label}", 1, 0, 'L', fill=True) 

                pdf_obj.set_font('Arial', '', 11)
                pdf_obj.multi_cell(130, 12, texto_completo, 1, 'L')
                pdf_obj.set_y(y_actual + altura_total)

            agregar_fila_dato(pdf, "UBICACIÃ“N", fila['Location'])
            agregar_fila_dato(pdf, "SUPERFICIE TOTAL", f"{fila['Area Final']:.2f}", "HectÃ¡reas")
            agregar_fila_dato(pdf, "INSUMO APLICADO", f"{fila['insumo_num']:.2f}", "L/Kg")
            agregar_fila_dato(pdf, "TIEMPO DE OPERACIÃ“N", formatear_tiempo(fila['segundos_vuelo']))
            agregar_fila_dato(pdf, "CANTIDAD DE VUELOS", int(fila['Flight time']))

            pdf_bytes = pdf.output(dest='S').encode('latin-1')
            # Nombre del archivo basado en fecha y ubicaciÃ³n abreviada
            loc_abreviada = str(fila['Location'])[:10].replace(" ", "_")
            nombre_pdf = f"Reporte_{fila['fecha_simple']}_{loc_abreviada}.pdf"
            zf.writestr(nombre_pdf, pdf_bytes)

    buffer_zip.seek(0)
    return buffer_zip.getvalue()

# --- APP PRINCIPAL ---
def main():
    st.title("ğŸŒ¾ AgroReport: Procesador de Operaciones")
    # --- BARRA LATERAL DE PERSONALIZACIÃ“N ---
    st.sidebar.header("PersonalizaciÃ³n del PDF")
    
    # Creamos el input. El nombre por defecto es "AGROFLY"
    nombre_empresa = st.sidebar.text_input("Nombre de la Empresa", value="AGROFLY")
    
    # Guardamos el nombre en la sesiÃ³n
    st.session_state['nombre_empresa'] = nombre_empresa.upper() # Lo pasamos a mayÃºsculas

    
    st.markdown("SubÃ­ el log de tu drone y elegÃ­ quÃ© reportes descargar.")

    uploaded_file = st.file_uploader("ElegÃ­ el archivo del drone (.xlsx)", type=['xlsx'])

    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
        informe = procesar_datos_informe(df)
        
        st.subheader("ğŸ“Š Reportes Detectados")
        st.write("SeleccionÃ¡ los lotes que querÃ©s procesar:")

        # 1. El usuario elige los reportes
        seleccion = st.multiselect(
            "Reportes disponibles:",
            options=informe['etiqueta'].tolist(),
            default=informe['etiqueta'].tolist(),
            help="HacÃ© clic para quitar o agregar reportes al paquete ZIP"
        )

        if seleccion:
            # Filtramos el dataframe de informe segÃºn la selecciÃ³n
            informe_final = informe[informe['etiqueta'].isin(seleccion)]
            
            st.info(f"Seleccionaste {len(informe_final)} reporte(s).")

            # 2. Generamos el ZIP solo con lo seleccionado
            if st.button("ğŸš€ Preparar Archivos para Descarga"):
                with st.spinner("Generando PDFs..."):
                    zip_data = generar_zip_seleccionado(informe_final)
                    
                    st.success("âœ… Â¡Paquete listo!")
                    st.download_button(
                        label="ğŸ“¥ Descargar Reportes Seleccionados (ZIP)",
                        data=zip_data,
                        file_name="Reportes_AgroFly_Seleccion.zip",
                        mime="application/zip"
                    )
        else:
            st.warning("âš ï¸ SeleccionÃ¡ al menos un reporte de la lista de arriba.")

if __name__ == "__main__":
    main()
